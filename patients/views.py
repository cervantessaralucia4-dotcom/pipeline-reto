# ═══════════════════════════════════════════════════════════════
#  patients/views.py
# ═══════════════════════════════════════════════════════════════
from drf_spectacular.utils import extend_schema
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from authentication.permissions import IsAdministrador, IsMedicoOrAnalista
from .models import Patient
from .serializers import PatientSerializer
from .predict import predict_risk


# ── PACIENTES CRUD ───────────────────────────────────────────
@extend_schema(tags=['Pacientes'])
class PatientViewSet(viewsets.ModelViewSet):
    """
    CRUD completo de pacientes clínicos.
    - GET    /api/patients/        — listar (Médico, Analista, Admin)
    - POST   /api/patients/        — crear  (solo Admin)
    - GET    /api/patients/{id}/   — detalle (Médico, Analista, Admin)
    - PUT    /api/patients/{id}/   — actualizar (solo Admin)
    - DELETE /api/patients/{id}/   — eliminar (solo Admin)
    """
    queryset         = Patient.objects.all()
    serializer_class = PatientSerializer

    def get_permissions(self):
        if self.action in ('list', 'retrieve'):
            return [IsMedicoOrAnalista()]
        return [IsAdministrador()]


# ── PREDICCIÓN ML ────────────────────────────────────────────
@extend_schema(
    tags=['Machine Learning'],
    summary='Predecir riesgo (acceso directo)',
    description='Alias de /api/ml/predict/. Acceso: Médico, Analista y Administrador.',
)
@api_view(['POST'])
@permission_classes([IsMedicoOrAnalista])
def predict_view(request):
    resultado = predict_risk(request.data)
    return Response({'prediction': resultado})


# ── DASHBOARD KPIs ───────────────────────────────────────────
@extend_schema(
    tags=['Dashboard'],
    summary='KPIs principales del dashboard',
    description='Total pacientes, críticos, riesgo alto/medio/bajo, glucosa e IMC promedio. Acceso: todos los roles.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_kpis(request):
    kpis = Patient.objects.aggregate(
        total=Count('id'),
        critico=Count('id', filter=Q(disease_risk='Crítico')),
        alto=Count('id', filter=Q(disease_risk='Alto')),
        medio=Count('id', filter=Q(disease_risk='Medio')),
        bajo=Count('id', filter=Q(disease_risk='Bajo')),
        avg_gluc=Avg('glucose'),
        avg_bmi=Avg('bmi'),
        avg_sys=Avg('systolic_pressure'),
        avg_dia=Avg('diastolic_pressure'),
        avg_hr=Avg('heart_rate'),
        avg_chol=Avg('cholesterol'),
        avg_o2=Avg('oxygen_saturation'),
        avg_temp=Avg('temperature'),
    )

    return Response({
        'total_patients':       kpis['total'] or 0,
        'critical_patients':    kpis['critico'] or 0,
        'high_risk':            kpis['alto'] or 0,
        'medium_risk':          kpis['medio'] or 0,
        'low_risk':             kpis['bajo'] or 0,
        'average_glucose':      round(kpis['avg_gluc'] or 0, 2),
        'average_bmi':          round(kpis['avg_bmi'] or 0, 2),
        'average_systolic':     round(kpis['avg_sys'] or 0, 1),
        'average_diastolic':    round(kpis['avg_dia'] or 0, 1),
        'average_heart_rate':   round(kpis['avg_hr'] or 0, 1),
        'average_cholesterol':  round(kpis['avg_chol'] or 0, 2),
        'average_oxygen_sat':   round(kpis['avg_o2'] or 0, 1),
        'average_temperature':  round(kpis['avg_temp'] or 0, 1),
    })


# ── DASHBOARD CHARTS ─────────────────────────────────────────
@extend_schema(
    tags=['Dashboard'],
    summary='Datos para gráficas del dashboard',
    description='Distribución por riesgo y sexo. Acceso: todos los roles.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_charts(request):
    dist = Patient.objects.aggregate(
        critico=Count('id', filter=Q(disease_risk='Crítico')),
        alto=Count('id', filter=Q(disease_risk='Alto')),
        medio=Count('id', filter=Q(disease_risk='Medio')),
        bajo=Count('id', filter=Q(disease_risk='Bajo')),
        masc=Count('id', filter=Q(sex='M')),
        fem=Count('id', filter=Q(sex='F'))
    )
    return Response({
        'risk_distribution': {
            'Crítico': dist['critico'] or 0,
            'Alto':    dist['alto'] or 0,
            'Medio':   dist['medio'] or 0,
            'Bajo':    dist['bajo'] or 0,
        },
        'gender_distribution': {
            'Masculino': dist['masc'] or 0,
            'Femenino':  dist['fem'] or 0,
        },
    })


# ── REPORTES ─────────────────────────────────────────────────
@extend_schema(
    tags=['Reportes'],
    summary='Reporte general del sistema',
    description='Resumen ejecutivo con totales, promedios y distribución de riesgo. Acceso: todos los roles.',
)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_view(request):
    stats = Patient.objects.aggregate(
        total=Count('id'),
        critico=Count('id', filter=Q(disease_risk='Crítico')),
        alto=Count('id', filter=Q(disease_risk='Alto')),
        medio=Count('id', filter=Q(disease_risk='Medio')),
        bajo=Count('id', filter=Q(disease_risk='Bajo')),
        avg_gluc=Avg('glucose'),
        avg_bmi=Avg('bmi'),
        avg_sys=Avg('systolic_pressure'),
        avg_dia=Avg('diastolic_pressure'),
        avg_hr=Avg('heart_rate'),
        avg_chol=Avg('cholesterol'),
        avg_o2=Avg('oxygen_saturation'),
        avg_temp=Avg('temperature'),
    )
    return Response({
        'total_patients':       stats['total'] or 0,
        'critical_patients':    stats['critico'] or 0,
        'average_glucose':      round(stats['avg_gluc'] or 0, 2),
        'average_bmi':          round(stats['avg_bmi'] or 0, 2),
        'average_systolic':     round(stats['avg_sys'] or 0, 1),
        'average_diastolic':    round(stats['avg_dia'] or 0, 1),
        'average_heart_rate':   round(stats['avg_hr'] or 0, 1),
        'average_cholesterol':  round(stats['avg_chol'] or 0, 2),
        'average_oxygen_sat':   round(stats['avg_o2'] or 0, 1),
        'average_temperature':  round(stats['avg_temp'] or 0, 1),
        'risk_distribution': {
            'Crítico': stats['critico'] or 0,
            'Alto':    stats['alto'] or 0,
            'Medio':   stats['medio'] or 0,
            'Bajo':    stats['bajo'] or 0,
        },
    })


# ── EXPORTAR CSV ──────────────────────────────────────────────
@extend_schema(
    tags=['Reportes'],
    summary='Exportar pacientes a CSV',
    description='Descarga un archivo CSV con todos los pacientes. Acceso: Médico, Analista y Administrador.',
)
@api_view(['GET'])
@permission_classes([IsMedicoOrAnalista])
def export_csv_view(request):
    campos = [
        'id', 'first_name', 'last_name', 'age', 'sex', 'weight', 'height', 'bmi',
        'systolic_pressure', 'diastolic_pressure', 'heart_rate', 'glucose',
        'cholesterol', 'oxygen_saturation', 'temperature', 'family_history',
        'smoker', 'alcohol_consumption', 'physical_activity',
        'preliminary_diagnosis', 'disease_risk', 'consultation_date',
    ]
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="pacientes.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    encabezados = ['ID', 'Nombre', 'Apellido', 'Edad', 'Sexo', 'Peso', 'Altura',
                   'IMC', 'Presión Sistólica', 'Presión Diastólica',
                   'Frecuencia Cardíaca', 'Glucosa', 'Colesterol',
                   'Saturación O₂', 'Temperatura', 'Antecedentes Familiares',
                   'Fumador', 'Consumo Alcohol', 'Actividad Física',
                   'Diagnóstico Preliminar', 'Riesgo', 'Fecha Consulta']
    writer.writerow(encabezados)
    for p in Patient.objects.all().values_list(*campos):
        writer.writerow(p)
    return response


# ── EXPORTAR EXCEL ─────────────────────────────────────────────
@extend_schema(
    tags=['Reportes'],
    summary='Exportar pacientes a Excel',
    description='Descarga un archivo Excel con todos los pacientes y hoja de resumen estadístico.',
)
@api_view(['GET'])
@permission_classes([IsMedicoOrAnalista])
def export_excel_view(request):
    pacientes = Patient.objects.all().order_by('id')
    wb = Workbook()

    # ── Hoja 1: Pacientes ──
    ws = wb.active
    ws.title = "Pacientes"

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['ID', 'Nombre', 'Apellido', 'Edad', 'Sexo', 'Peso', 'Altura',
               'IMC', 'Presión Sistólica', 'Presión Diastólica',
               'Frecuencia Cardíaca', 'Glucosa', 'Colesterol',
               'Saturación O₂', 'Temperatura', 'Antecedentes Familiares',
               'Fumador', 'Consumo Alcohol', 'Actividad Física',
               'Diagnóstico Preliminar', 'Riesgo', 'Fecha Consulta']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, p in enumerate(pacientes, 2):
        values = [p.id, p.first_name, p.last_name, p.age, p.sex, p.weight, p.height,
                  p.bmi, p.systolic_pressure, p.diastolic_pressure,
                  p.heart_rate, p.glucose, p.cholesterol,
                  p.oxygen_saturation, p.temperature,
                  'Sí' if p.family_history else 'No',
                  'Sí' if p.smoker else 'No',
                  'Sí' if p.alcohol_consumption else 'No',
                  p.physical_activity, p.preliminary_diagnosis,
                  p.disease_risk, p.consultation_date]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['T'].width = 28
    ws.auto_filter.ref = ws.dimensions

    # ── Hoja 2: Resumen ──
    ws2 = wb.create_sheet("Resumen Estadístico")
    stats = Patient.objects.aggregate(
        total=Count('id'),
        avg_edad=Avg('age'), avg_peso=Avg('weight'), avg_altura=Avg('height'),
        avg_bmi=Avg('bmi'), avg_sys=Avg('systolic_pressure'),
        avg_dia=Avg('diastolic_pressure'), avg_hr=Avg('heart_rate'),
        avg_gluc=Avg('glucose'), avg_chol=Avg('cholesterol'),
        avg_o2=Avg('oxygen_saturation'), avg_temp=Avg('temperature'),
    )
    resumen = [
        ('Métrica', 'Valor'),
        ('Total Pacientes', stats['total'] or 0),
        ('Edad Promedio', round(stats['avg_edad'] or 0, 1)),
        ('Peso Promedio (kg)', round(stats['avg_peso'] or 0, 2)),
        ('Altura Promedio (m)', round(stats['avg_altura'] or 0, 2)),
        ('IMC Promedio', round(stats['avg_bmi'] or 0, 2)),
        ('Presión Sistólica Promedio', round(stats['avg_sys'] or 0, 1)),
        ('Presión Diastólica Promedio', round(stats['avg_dia'] or 0, 1)),
        ('Frecuencia Cardíaca Promedio', round(stats['avg_hr'] or 0, 1)),
        ('Glucosa Promedio (mg/dL)', round(stats['avg_gluc'] or 0, 2)),
        ('Colesterol Promedio (mg/dL)', round(stats['avg_chol'] or 0, 2)),
        ('Saturación O₂ Promedio (%)', round(stats['avg_o2'] or 0, 1)),
        ('Temperatura Promedio (°C)', round(stats['avg_temp'] or 0, 1)),
    ]
    for row_idx, (label, val) in enumerate(resumen, 1):
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_b = ws2.cell(row=row_idx, column=2, value=val)
        cell_a.font = Font(name='Calibri', bold=(row_idx == 1), color='FFFFFF' if row_idx == 1 else '000000')
        cell_a.fill = header_fill if row_idx == 1 else PatternFill()
        cell_b.font = Font(name='Calibri', bold=(row_idx == 1), color='FFFFFF' if row_idx == 1 else '000000')
        cell_b.fill = header_fill if row_idx == 1 else PatternFill()
        cell_a.alignment = header_align if row_idx == 1 else Alignment(horizontal='left')
        cell_b.alignment = header_align
        cell_a.border = thin_border
        cell_b.border = thin_border

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 22

    # ── Hoja 3: Distribución ──
    ws3 = wb.create_sheet("Distribución por Riesgo")
    riesgo_data = Patient.objects.values('disease_risk').annotate(
        cantidad=Count('id'), avg_gluc=Avg('glucose'), avg_bmi=Avg('bmi')
    ).order_by('-cantidad')
    headers_r = ['Nivel de Riesgo', 'Cantidad', 'Glucosa Promedio', 'IMC Promedio']
    for col, h in enumerate(headers_r, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin_border
    for row_idx, r in enumerate(riesgo_data, 2):
        for col, val in enumerate([r['disease_risk'], r['cantidad'],
                                    round(r['avg_gluc'] or 0, 2),
                                    round(r['avg_bmi'] or 0, 2)], 1):
            cell = ws3.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
    for col in range(1, 5):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_pacientes.xlsx"'
    return response


# ── EXPORTAR PDF ──────────────────────────────────────────────
@extend_schema(
    tags=['Reportes'],
    summary='Exportar pacientes a PDF',
    description='Descarga un reporte PDF con todos los pacientes y resumen estadístico.',
)
@api_view(['GET'])
@permission_classes([IsMedicoOrAnalista])
def export_pdf_view(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.units import mm, cm
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    import textwrap

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                  fontSize=18, spaceAfter=6, textColor=colors.HexColor('#1F4E79'))
    elements.append(Paragraph('Reporte General de Pacientes', title_style))
    elements.append(Spacer(1, 4*mm))

    # Resumen
    stats = Patient.objects.aggregate(
        total=Count('id'),
        avg_gluc=Avg('glucose'), avg_bmi=Avg('bmi'),
        avg_sys=Avg('systolic_pressure'), avg_dia=Avg('diastolic_pressure'),
        avg_hr=Avg('heart_rate'), avg_temp=Avg('temperature'),
    )
    resumen = [
        ['Total Pacientes', str(stats['total'] or 0)],
        ['Glucosa Promedio', f"{round(stats['avg_gluc'] or 0, 2)} mg/dL"],
        ['IMC Promedio', str(round(stats['avg_bmi'] or 0, 2))],
        ['Presión Arterial Prom', f"{round(stats['avg_sys'] or 0, 1)}/{round(stats['avg_dia'] or 0, 1)} mmHg"],
        ['Frec. Cardíaca Prom', f"{round(stats['avg_hr'] or 0, 1)} lpm"],
        ['Temperatura Prom', f"{round(stats['avg_temp'] or 0, 1)} °C"],
    ]
    t_resumen = Table(resumen, colWidths=[120, 100])
    t_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F7FB')]),
    ]))
    elements.append(t_resumen)
    elements.append(Spacer(1, 6*mm))

    # Tabla de pacientes
    pacientes = Patient.objects.all().order_by('id')
    data = [['ID', 'Nombre', 'Edad', 'Sexo', 'P. Sist', 'P. Diast',
             'FC', 'Glucosa', 'Colesterol', 'Sat O₂', 'Temp', 'IMC', 'Riesgo']]
    for p in pacientes:
        data.append([
            str(p.id), f"{p.first_name} {p.last_name}", str(p.age),
            p.sex, str(p.systolic_pressure or ''), str(p.diastolic_pressure or ''),
            str(p.heart_rate or ''), str(p.glucose or ''), str(p.cholesterol or ''),
            str(p.oxygen_saturation or ''), str(p.temperature or ''),
            str(p.bmi or ''), p.disease_risk
        ])

    col_widths = [25, 80, 25, 25, 30, 30, 28, 35, 35, 30, 28, 28, 40]
    t_pacientes = Table(data, colWidths=col_widths, repeatRows=1)
    t_pacientes.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F7FB')]),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
    ]))
    elements.append(t_pacientes)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_pacientes.pdf"'
    return response